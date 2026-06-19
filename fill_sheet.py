"""
Read NACE codes out of the TAM workbook, pull Eurostat counts, and write the
SME / Corp(250+) columns back in.

Two hard facts about the data drive the design:

1. `sbs_sc_ovw` has NO 4-digit class data — only section (C), division (C27),
   and group (C279). The sheet is full of class codes (C26.30, D35.11, ...), so
   every code is resolved DOWN a fallback chain: class -> group -> division ->
   section, using the most specific level that actually exists, and the level
   used is reported per row.

2. Several class codes in one segment collapse to the same group/division
   (e.g. C33.15 + C33.16 -> C331). Those are de-duplicated so a firm population
   is never summed twice within a row.

Only YELLOW input cells (F = SMEs, G = Corp Mid-cap 250+) on segment rows are
written. Blue subtotals, the grey public-sector row, and demand-side rows are
left untouched. Output goes to a COPY by default — the original is never edited
in place unless you explicitly ask for it.
"""

from __future__ import annotations

import re
from typing import Any

import openpyxl

import server  # reuse _fetch / _decode and the bucket constants

SHEET = "Company Counts by Segment"
COL_SEGMENT = 2   # B
COL_NACE = 3      # C
COL_SME = 6       # F  -> SMEs
COL_CORP = 7      # G  -> Corp: Mid-cap (250-4,999)  [= all 250+ from Eurostat]
YELLOW = "FFFFF2CC"   # input-cell fill used in the sheet

# Phrases that mark a row as not a Eurostat company population.
SKIP_PHRASES = (
    "demand side", "demand-side", "outside sbs", "cross-vertical",
    "sector-wide", "not a company population",
)

_CODE = re.compile(r"([A-Z])(\d{2})(?:\.(\d{1,2}))?")   # lettered NACE code
_CONT = re.compile(r"[/\-]\s*(\d{2})(?:\.(\d{1,2}))?")   # continuation: /35.14 or -35.23


# --------------------------------------------------------------------------- #
# NACE parsing
# --------------------------------------------------------------------------- #
def parse_nace(text: str) -> list[str]:
    """
    Extract dotted NACE codes from the sheet's free-text cell.

    Handles 'C27', 'D35.11', slash continuations ('C21.10/21.20' -> both,
    'D35.11/35.14' -> both inheriting 'D'), and range endpoints
    ('D35.22-35.23' -> both endpoints). Bare letterless numbers that are NOT
    slash/range continuations (e.g. a parenthetical '(28.11 turbines)') are
    ignored, because they merely annotate a division already captured.
    """
    if not text:
        return []
    found: list[str] = []
    # 1) lettered codes, remembering each match's letter for nearby continuations
    matches = list(_CODE.finditer(text))
    for m in matches:
        letter, div, sub = m.group(1), m.group(2), m.group(3)
        found.append(f"{letter}{div}" + (f".{sub}" if sub else ""))
    # 2) continuations (/NN.NN or -NN.NN) inherit the letter of the preceding code
    for cm in _CONT.finditer(text):
        # nearest lettered code starting before this continuation
        prev = [m for m in matches if m.start() < cm.start()]
        if not prev:
            continue
        letter = prev[-1].group(1)
        div, sub = cm.group(1), cm.group(2)
        found.append(f"{letter}{div}" + (f".{sub}" if sub else ""))
    # de-dup, preserve order
    seen: set[str] = set()
    out = []
    for c in found:
        if c not in seen:
            seen.add(c)
            out.append(c)
    return out


def resolve(code: str, available: set[str]) -> tuple[str | None, str]:
    """Walk class -> group -> division -> section, return first level present."""
    p = code.replace(".", "").upper()
    letter, digits = p[0], p[1:]
    for n in (len(digits), 3, 2):
        if 0 < n <= len(digits):
            cand = letter + digits[:n]
            if cand in available:
                return cand, {4: "class", 3: "group", 2: "division"}[n]
    if letter in available:
        return letter, "section"
    return None, "none"


# --------------------------------------------------------------------------- #
# Eurostat: codelist + batched counts for resolved codes
# --------------------------------------------------------------------------- #
def _codelist(geo: str = "DE") -> set[str]:
    data = server._fetch(
        [("indic_sbs", "ENT_NR"), ("geo", geo), ("size_emp", "TOTAL"), ("time", "2022")]
    )
    return set(data["dimension"]["nace_r2"]["category"]["index"].keys())


def _counts(codes: list[str], geo: list[str], year: str | None,
            exclude_micro: bool) -> dict[str, dict[str, Any]]:
    """One batched pull; returns {resolved_code: {sme, corp, year}}."""
    if not codes:
        return {}
    sme_bands = server.SME_BANDS_NO_MICRO if exclude_micro else server.SME_BANDS
    params: list[tuple[str, str]] = [("indic_sbs", "ENT_NR")]
    for c in codes:
        params.append(("nace_r2", c))
    for g in geo:
        params.append(("geo", g))
    for s in sme_bands + [server.CORP_BAND]:
        params.append(("size_emp", s))
    if year:
        params.append(("time", year))
    rows = server._decode(server._fetch(params))

    out: dict[str, dict[str, Any]] = {}
    for code in codes:
        crows = [r for r in rows if r["nace_r2"] == code]
        if not crows:
            out[code] = {"sme": 0, "corp": 0, "year": None, "missing": True}
            continue
        yr = year or max(r["time"] for r in crows)
        crows = [r for r in crows if r["time"] == yr]
        out[code] = {
            "sme": sum(r["value"] for r in crows if r["size_emp"] in sme_bands),
            "corp": sum(r["value"] for r in crows if r["size_emp"] == server.CORP_BAND),
            "year": yr,
            "missing": False,
        }
    return out


# --------------------------------------------------------------------------- #
# Plan: what each row would get (no file write)
# --------------------------------------------------------------------------- #
def compute_plan(xlsx_path: str, geo: list[str] | None = None,
                 year: str | None = None, exclude_micro: bool = False) -> dict[str, Any]:
    geo = geo or server.DEFAULT_GEO
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    ws = wb[SHEET]
    available = _codelist()

    plan_rows: list[dict[str, Any]] = []
    all_resolved: set[str] = set()

    for r in range(22, ws.max_row + 1):
        cell = ws.cell(r, COL_SME)
        is_yellow = bool(cell.fill and cell.fill.patternType
                         and cell.fill.fgColor.rgb == YELLOW)
        if not is_yellow:
            continue
        nace_text = ws.cell(r, COL_NACE).value or ""
        segment = ws.cell(r, COL_SEGMENT).value or ""
        lowered = nace_text.lower()
        if any(p in lowered for p in SKIP_PHRASES):
            plan_rows.append({"row": r, "segment": segment, "status": "skipped-demand-side",
                              "nace_text": nace_text, "resolved": []})
            continue
        parsed = parse_nace(nace_text)
        resolved: list[dict[str, str]] = []
        seen: set[str] = set()
        for code in parsed:
            rc, lvl = resolve(code, available)
            if rc is None:
                resolved.append({"requested": code, "resolved": None, "level": "none"})
                continue
            if rc in seen:
                resolved.append({"requested": code, "resolved": rc, "level": lvl, "deduped": True})
                continue
            seen.add(rc)
            all_resolved.add(rc)
            resolved.append({"requested": code, "resolved": rc, "level": lvl})
        plan_rows.append({"row": r, "segment": segment, "status": "ok" if parsed else "no-codes",
                          "nace_text": nace_text, "resolved": resolved})

    counts = _counts(sorted(all_resolved), geo, year, exclude_micro)

    # assemble per-row totals from unique resolved codes
    for pr in plan_rows:
        if pr["status"] != "ok":
            pr["sme"] = pr["corp"] = None
            continue
        used = [x["resolved"] for x in pr["resolved"]
                if x.get("resolved") and not x.get("deduped")]
        pr["sme"] = sum(counts[c]["sme"] for c in used)
        pr["corp"] = sum(counts[c]["corp"] for c in used)
        pr["years"] = sorted({counts[c]["year"] for c in used if counts[c]["year"]})
        pr["coarsened"] = sorted({x["resolved"] for x in pr["resolved"]
                                  if x["level"] in ("division", "section")})

    return {"geo": geo, "exclude_micro": exclude_micro, "rows": plan_rows,
            "codes_pulled": len(all_resolved)}


# --------------------------------------------------------------------------- #
# Write
# --------------------------------------------------------------------------- #
def fill_workbook(input_path: str, output_path: str, geo: list[str] | None = None,
                  year: str | None = None, exclude_micro: bool = False) -> dict[str, Any]:
    plan = compute_plan(input_path, geo, year, exclude_micro)
    wb = openpyxl.load_workbook(input_path, data_only=False)
    ws = wb[SHEET]
    written = 0
    for pr in plan["rows"]:
        if pr["status"] != "ok":
            continue
        ws.cell(pr["row"], COL_SME).value = pr["sme"]
        ws.cell(pr["row"], COL_CORP).value = pr["corp"]
        written += 1
    wb.save(output_path)
    return {"output_path": output_path, "rows_written": written,
            "rows_skipped": sum(1 for p in plan["rows"] if p["status"] != "ok"),
            "plan": plan}
